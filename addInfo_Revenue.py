# -*- coding: utf-8 -*-

import json
import os

class addInfo_Revenue:
    def set_read_only_revenue(self):
        for w in [self.lne_tong1, self.lne_phikhac15, self.lne_tongphiwifi1,
                  self.lne_ttiennuoc1, self.lne_ttiendien1,self.lne_ttienphong1,self.lne_thangnam6]:
            w.setReadOnly(True)
            
    def report(self):
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
        from PyQt6.QtWidgets import QFileDialog, QMessageBox

    
        # ── Đọc dữ liệu từ các ô trên form =========================================================
        thangnam   = self.lne_thangnam6.text().strip()
        tienphong  = self.lne_ttienphong1.text().strip()
        tiendien   = self.lne_ttiendien1.text().strip()
        tiennuoc   = self.lne_ttiennuoc1.text().strip()
        phivesinh  = self.lne_tphivs1.text().strip()
        phiguixe   = self.lne_tphigiuxe1.text().strip()
        phiwifi    = self.lne_tongphiwifi1.text().strip()
        phikhac    = self.lne_phikhac15.text().strip()
        tong       = self.lne_tong1.text().strip()
    
        # ── Chọn nơi lưu file ==========================================================================
        default_name = f"BaoCaoDoanhThu_{thangnam.replace('/', '-')}.xlsx"
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Lưu báo cáo doanh thu", default_name,
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return
    
        # ── Tạo workbook=================================================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Doanh Thu"
    
        # Helper styles
        thin = Side(style="thin", color="BFBFBF")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    
        def hdr_cell(row, col, value, bg="1E3A5F", fg="FFFFFF", size=11, bold=True):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="Arial", bold=bold, size=size, color=fg)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border_all
            return c
    
        def data_cell(row, col, value, bold=False, align="right", bg="FFFFFF"):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="Arial", size=11, bold=bold)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = border_all
            return c
    
        # ── Tiêu đề chính ===================================================================================
        ws.merge_cells("A1:C1")
        title = ws["A1"]
        title.value = f"BÁO CÁO DOANH THU  —  Tháng {thangnam}"
        title.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill("solid", start_color="1E3A5F")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
    
        ws.merge_cells("A2:C2")
        sub = ws["A2"]
        sub.font = Font(name="Arial", italic=True, size=10, color="595959")
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
    
        # ── Header bảng ===================================================================================
        ws.row_dimensions[3].height = 28
        hdr_cell(3, 1, "Khoản thu")
        hdr_cell(3, 2, "Số tiền (VNĐ)")
        hdr_cell(3, 3, "Ghi chú")
    
        # ── Dữ liệu===================================================================================
        rows = [
            ("Tiền phòng",    tienphong),
            ("Tiền điện",     tiendien),
            ("Tiền nước",     tiennuoc),
            ("Phí vệ sinh",   phivesinh),
            ("Phí gửi xe",    phiguixe),
            ("Phí wifi",      phiwifi),
            ("Phí khác",      phikhac),
        ]
    
        row_colors = ["F7F9FC", "FFFFFF"]
        for i, (label, value) in enumerate(rows, start=4):
            bg = row_colors[i % 2]
            ws.row_dimensions[i].height = 24
            data_cell(i, 1, label, align="left",  bg=bg)
            data_cell(i, 2, value, align="right", bg=bg)
            data_cell(i, 3, "",   align="left",   bg=bg)
    
        # ── Dòng TỔNG ===================================================================================
        total_row = 4 + len(rows)
        ws.row_dimensions[total_row].height = 28
        hdr_cell(total_row, 1, "TỔNG DOANH THU", bg="2E6DA4", size=11)
        hdr_cell(total_row, 2, tong,             bg="2E6DA4", size=11)
        hdr_cell(total_row, 3, "",               bg="2E6DA4", size=11)
    
        # ── Độ rộng cột===================================================================================
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 24
    
        # ── Lưu file===================================================================================
        try:
            wb.save(filepath)
            QMessageBox.information(
                self, "Thành công",
                f"Đã xuất báo cáo doanh thu thành công!\n{filepath}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không lưu được file:\n{e}")